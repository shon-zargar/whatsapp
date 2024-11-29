import time
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import logging
import os
from openpyxl import load_workbook, Workbook
from io import StringIO
from pytest_html import extras
import pytest
import conftest

# Configure logging
log_stream = StringIO()
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s', stream=log_stream)

# Chrome profile directory path
chrome_profile_directory = "C:/Users/shonz/AppData/Local/Google/Chrome/User Data/Default"

# Fixture to set up the browser
@pytest.fixture(scope="class")
def setup(request):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f"user-data-dir={chrome_profile_directory}")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--ignore-certificate-errors")
    service_obj = Service("C:/Users/shonz/Desktop/selenium/chromedriver2/chromedriver.exe")
    driver = webdriver.Chrome(service=service_obj, options=chrome_options)
    driver.implicitly_wait(5)
    driver.get("https://web.whatsapp.com/")
    time.sleep(10)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "(//button[@type='button'])[2]")))
    request.cls.driver = driver
    yield driver
    driver.quit()

@pytest.mark.usefixtures("setup")
class TestWhatsApp:

    @pytest.mark.parametrize("term", conftest.terms)
    def test_check_text(self, term, request):
        request.node.name = f"test_check_text_{term}"
        time.sleep(40)
        wait = WebDriverWait(self.driver, 10)
        logging.info("Waiting for the button element to be clickable.")
        not_called = wait.until(EC.presence_of_element_located((By.XPATH, "(//button[@type='button'])[2]")))
        assert not_called is not None, "Element not found!"
        if not_called.get_attribute("aria-pressed") == "false":
            logging.info("Element is clickable. Clicking on it.")
            not_called.click()
        else:
            logging.info("Element is already clicked.")

        screenshot_path = f'C:/Users/shonz/PycharmProjects/pythonProject/tests/screenshots/screenshot_{term}.png'
        search_field = None

        try:
            logging.info(f"Starting search for term: {term}")
            search_field = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '.selectable-text.copyable-text.x15bjb6t.x1n2onr6'))
            )
            search_field.click()
            search_field.send_keys(term)
            time.sleep(6)
            logging.info(f"Search for term '{term}' completed successfully.")
            WebDriverWait(self.driver, 5).until(
                EC.presence_of_all_elements_located((By.XPATH, "//div[@role='listitem']"))
            )

            div_elements = self.driver.find_elements(By.XPATH, "//div[@role='listitem']")
            os.makedirs(os.path.dirname(screenshot_path), exist_ok=True)
            self.driver.save_screenshot(screenshot_path)
            logging.info(f"Screenshot saved at {screenshot_path}")
            self.add_screenshot_to_report(request, screenshot_path)

            if div_elements:
                logging.info(f"Results found for term: {term}. Saving to Excel file.")

                excel_path = f'C:/Users/shonz/PycharmProjects/{term}_Results.xlsx'
                sheet_name = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

                # Create a new Excel file if it doesn't exist
                if not os.path.exists(excel_path):
                    logging.info(f"Creating new Excel file: {excel_path}")
                    book = Workbook()
                    book.save(excel_path)

                # Load existing data from the Excel file
                existing_data = set()
                book = load_workbook(excel_path)
                for sheet in book.sheetnames:
                    df_existing = pd.read_excel(excel_path, sheet_name=sheet)
                    existing_data.update(df_existing.apply(tuple, axis=1))

                # Create a new list of unique data
                div_list = [div.text.split("\n") for div in div_elements]
                df_new = pd.DataFrame(div_list).dropna()
                new_data = [tuple(row) for row in df_new.values if tuple(row) not in existing_data]

                # Save unique data
                if new_data:
                    df_unique = pd.DataFrame(new_data)
                    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
                        df_unique.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
                    logging.info(f"New data saved in {excel_path} under sheet {sheet_name}.")
                else:
                    logging.info("No new data to save.")

            self.clear_search_field(search_field)
            logging.info("Cleared previous text from the search field.")

        except Exception as e:
            logging.warning(f"No results found for term: {term}. Error: {str(e)}")
            if search_field:
                self.clear_search_field(search_field)
            self.add_screenshot_to_report(request, screenshot_path)

    def clear_search_field(self, search_field):
        logging.info("Clearing the search field.")
        search_field.click()
        search_field.send_keys(Keys.CONTROL + "a")
        search_field.send_keys(Keys.BACKSPACE)
        time.sleep(6)

    def add_screenshot_to_report(self, request, screenshot_path):
        try:
            if not os.path.exists(screenshot_path):
                logging.warning(f"Screenshot not found at: {screenshot_path}")
                return

            extra_image = extras.html(
                f'<a href="{screenshot_path}" target="_blank">'
                f'<img src="{screenshot_path}" alt="screenshot" style="max-width:300px; max-height:150px;"/></a>'
            )

            if not hasattr(request.node, 'extra'):
                request.node.extra = []
            request.node.extra.append(extra_image)

            logging.info("Screenshot added to report with clickable link.")
        except Exception as e:
            logging.error(f"Failed to add screenshot to report: {str(e)}")

    @pytest.hookimpl(tryfirst=True)
    def pytest_runtest_makereport(self, item, call):
        if call.when == "call":
            log_output = log_stream.getvalue()
            if log_output:
                extra_logs = extras.html(f"<div><pre>{log_output}</pre></div>")
                item.extra = getattr(item, 'extra', []) + [extra_logs]
            log_stream.truncate(0)
            log_stream.seek(0)
